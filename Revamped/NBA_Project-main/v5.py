from openpyxl.styles import Font, Alignment

from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl

import openpyxl                                                                          #import openpyxl module

from openpyxl import Workbook                                                         #import workbook from openpyxl

from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl

from openpyxl.styles.borders import Border, Side                                #import border from openpyxl

from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl

from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl

from openpyxl import load_workbook

import time                                                                                  #import time module

import numpy as np                                                                    #import numpy module

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder  #import column dimension from openpyxl

#import formula from openpyxl

from openpyxl.utils import FORMULAE

def adjust_width(ws):

    #adjust width of the columns in the worksheet including the merged cells

    for col in ws.columns:

        max_length = 0

        column = col[0].column

        for cell in col:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(cell.value)

            except:

                pass

            cell.alignment = Alignment(horizontal='center', vertical='center')

        adjusted_width = (max_length + 2) * 1.7

        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

#Input details

def input_detail(data,aw):  #function to input details

    i=1

    aw[f'A{i}']="Heading"

    aw[f'B{i}']="Inputs"

    i+=1

    for key, value in data.items():

        aw[f'A{i}']=key

        aw[f'B{i}']=value

        i+=1

    for row in aw.iter_rows(min_row=1, max_col=2, max_row=len(data.keys())+1):

        for cell in row:

            cell.font = Font(bold=True)

                                                                                                                         #add table to workshee

    tab=Table(displayName=f"Input_Details", ref=f"A1:B{len(data.keys())+1}")                                                                                 #create table

    style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)  #set style

    tab.tableStyleInfo = style                                                                                                                   #set style

    aw.add_table(tab)                                                                                                                            #add table to worksheet

    return aw                                                                                                                                    #return worksheet

def CO_PO_Table(data,aw):

    #merge cells depending on number of POs

    aw.merge_cells(start_row=1, start_column=4, end_row=1, end_column=12+5+1+3)

    aw['D1']="CO-PO Mapping"

    aw['D1'].font = Font(bold=True)

    aw['D1'].alignment = Alignment(horizontal='center', vertical='center')

    aw["D2"]="COs\POs"

    aw["D2"].font = Font(bold=True)

    for co in range(1,data["Number_of_COs"]+1):

        aw[f"D{co+2}"]=f"CO{co}"

        aw[f"D{co+2}"].font = Font(bold=True)

    for po in range(1,12+1):

        aw[f"{get_column_letter(po+4)}2"]=f"PO{po}   "

        aw[f"{get_column_letter(po+4)}2"].font = Font(bold=True)

    for pso in range(1,6):

        aw[f"{get_column_letter(12+4+pso)}2"]=f"PSO{pso}"

        aw[f"{get_column_letter(12+4+pso)}2"].font = Font(bold=True)

    #make it into a table

    tab = Table(displayName=f"CO_PO", ref=f"D2:{get_column_letter(12+4+5)}{data['Number_of_COs']+2}")  #create table

    style = TableStyleInfo(name="TableStyleMedium3", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)  #set style

    tab.tableStyleInfo = style                                                                                                                   #set style

    aw.add_table(tab)

    return aw

def indirect_co_assessment(data,aw):

    #merge cells depending on number of POs

    aw.merge_cells(start_row=data["Number_of_COs"]+5, start_column=4, end_row=data["Number_of_COs"]+5, end_column=5)

    aw[f'D{data["Number_of_COs"]+5}']="Indirect CO Assessment"

    aw[f'D{data["Number_of_COs"]+5}'].font = Font(bold=True)

    aw[f"D{data['Number_of_COs']+6}"]="COs\Components"

    aw[f"D{data['Number_of_COs']+6}"].font = Font(bold=True)

    # for let in range(1,5):

    #     aw[f"{get_column_letter(let+4)}{data['Number_of_COs']+6}"]=get_column_letter(let)

    #     aw[f"{get_column_letter(let+4)}{data['Number_of_COs']+6}"].font = Font(bold=True)

    for i in range(1,data["Number_of_COs"]+1):

        aw[f"D{i+data['Number_of_COs']+6}"]=f"CO{i}"

        aw[f"D{i+data['Number_of_COs']+6}"].font = Font(bold=True)

    aw[f"E{data['Number_of_COs']+6}"]="Component"

    aw[f"E{data['Number_of_COs']+6}"].font = Font(bold=True)

    #make it into a table

    tab = Table(displayName=f"Indirect_CO_Assessment", ref=f"D{data['Number_of_COs']+6}:E{data['Number_of_COs']+data['Number_of_COs']+6}")  #create table

    style = TableStyleInfo(name="TableStyleMedium14", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)  #set style

    tab.tableStyleInfo = style                                                                                                                   #set style

    aw.add_table(tab)

    return aw

#qn_co_mm_btl table

def qn_co_mm_btl(data,key, Component_details,aw):  #function to create qn_co_mm_btl table

    aw.merge_cells(f'B1:{get_column_letter(Component_details["Number_of_questions"]+2)}1')

    aw[f'B1']=key

    aw['B2']="Question"

    aw['B3']="Max Marks"

    aw['B4']="Threshold"

    aw['B5']="CO"

    aw['B6']="Final CO"

    aw['B7']="BTL"

    #Make them bold

    for row in aw.iter_rows(min_row=1, max_row=7, min_col=2, max_col=2):   #set font and alignment for heading

        for cell in row:                                                    #set font and alignment for heading

            cell.font = Font(bold=True)                                     #set font to bold

    for qno in range(1,Component_details['Number_of_questions']+1):

        aw[get_column_letter(qno+2)+'2']=f"Q{qno}"        

        aw[get_column_letter(qno+2) + '6'].value = f'=CONCATENATE("{data["Subject_Code"]+"_CO"}", {get_column_letter(qno+2)}5)'

        #aw[get_column_letter(qno+2)+'4']=f'={float(float(data["Default threshold %"])/100)}*{get_column_letter(qno+2)}3)'

        #aw[get_column_letter(qno+2)+'4']=f'={float(50/100)}*{get_column_letter(qno+2)}3)'

        aw[get_column_letter(qno+2)+'4']=f'={float(data["Default threshold %"])/100}*{get_column_letter(qno+2)}3'

    table_range = f"B2:{get_column_letter(Component_details['Number_of_questions'] + 2)}7"

    tab = Table(displayName=f"qn_co_mm_btl_{key}", ref=table_range)

    style = TableStyleInfo(name="TableStyleLight10", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style

    aw.add_table(tab)

    return aw

def studentmarks(data,key, Component_details,aw):

    aw.merge_cells(f'B9:{get_column_letter(Component_details["Number_of_questions"]+2)}9')

    aw["B9"]="Marks obtained"

    aw["B9"].font = Font(bold=True)

    aw["A10"]="Roll No."

    aw["A10"].font = Font(bold=True)

    aw["B10"]="Name"

    aw["B10"].font = Font(bold=True)

    for qno in range(1,Component_details['Number_of_questions']+1):

        aw[get_column_letter(qno+2)+'10']=f"Q{qno}"

        aw[get_column_letter(qno+2)+'10'].font = Font(bold=True)

    table_range = f"A10:{get_column_letter(Component_details['Number_of_questions'] + 2)}{data['Number_of_Students'] + 10}"

    tab = Table(displayName=f"studentmarks_{key}", ref=table_range)

    style = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style

    aw.add_table(tab)

    return aw

def main1(data,Component_Details):

    #create openpyxl workbook

    wb = Workbook()

    wb.remove(wb.active)

    # data={

    #     "Bundle_Number":"17-11-24-003",                                                             #set bundle number

    #     "Teacher":"Dr. S. S. Patil",                                                              #set teacher name

    #     "Academic_year":"2022-2023",                                                              #set academic year

    #     "Semester":7,                                                                                 #set semester

    #     "Branch":"CSE",                                                                          #set branch

    #     "Batch":2019,                                                                             #set batch

    #     "Section":"A",                                                                           #set section

    #     "Subject_Code":"19CSE345",                                                            #set subject code

    #     "Subject_Name":"Computer system and architecture",                          #set subject name

    #     "Number_of_Students":52,

    #     "Number_of_COs":6,

    #     "Internal":70,

    #     "External":30,

    #     "Direct":80,

    #     "Indirect":20,

    #     "Default threshold %":70                                                  

    #         }

    wb.create_sheet("Input Details")

    ws = wb["Input Details"]

    ws = input_detail(data,ws)

    ws = CO_PO_Table(data,ws)

    ws = indirect_co_assessment(data,ws)

    adjust_width(ws)

    # Component_Details={"P1_I":{"Number_of_questions":3},

    #                     "P2_I":{"Number_of_questions":6},

    #                     "CA_I":{"Number_of_questions":6},

    #                     "EndSem_E":{"Number_of_questions":9}}

    #get number of components

    number_of_components = len(Component_Details.keys())

    #iterate throught Keys of Component_Details and make a worksheet for each key

    for key in Component_Details.keys():

        wb.create_sheet(key)

        ws = wb[key]

        ws.title = key

        ws = qn_co_mm_btl(data, key, Component_Details[key], ws)

        ws = studentmarks(data, key, Component_Details[key], ws)   

        adjust_width(ws)

    #save workbook

    wb.save(f"{data['Batch']}_{data['Subject_Code']}_{data['Subject_Name']}.xlsx")
