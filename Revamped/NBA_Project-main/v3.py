
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
import openpyxl                                                                          #import openpyxl module
from openpyxl import Workbook                                                         #import workbook from openpyxl
from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.utils import get_column_letter, range_boundaries                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl import load_workbook
import time                                                                                  #import time module
import numpy as np                                                                    #import numpy module
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder  #import column dimension from openpyxl
#import max column
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string


import pandas as pd

def adjust_width(ws):
    #adjust width of the columns in the worksheet including the merged cells
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.5
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

def create_sheets(filename, wb):
    wb1 = load_workbook(filename, data_only=True)
    #get all sheet names
    sheetnames = wb1.sheetnames
    #drop sheetname 'Sheet' if it exists
    if "Input Details" in sheetnames:
        sheetnames.remove('Input Details')
    #set all the sheets to wb
    for sheetname in sheetnames:
        wb.create_sheet("CO_"+sheetname)
        
    return wb

#Calculate thhe cummulative CO
def CummulativeCO(awwrite,awread,table_colour):
    table1 = awread.tables[f'qn_co_mm_btl_{awread.title}']
    table_range1 = table1.ref
    min_col, min_row, max_col, max_row = range_boundaries(table_range1)
    # Create a list to store the table's data
    table_data1 = []
    # Iterate through the table rows and columns
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell_value = awread.cell(row=row, column=col).value
            row_data.append(cell_value)
        table_data1.append(row_data)
    #make it into a dataframe
    qno_mm_co_btl = pd.DataFrame(table_data1[1:], columns=table_data1[0])
    

    for i in range(1, len(qno_mm_co_btl.columns)):
        cell_value = qno_mm_co_btl.iloc[1, i]
        if cell_value is not None:
            cell_value = float(cell_value)
        else:
            cell_value = 0.5 * float(qno_mm_co_btl.iloc[0, i])
        qno_mm_co_btl.iloc[1, i] = cell_value

    #print(qno_mm_co_btl)

    table2 = awread.tables[f'studentmarks_{awread.title}']
    table_range2 = table2.ref
    min_col, min_row, max_col, max_row = range_boundaries(table_range2)
    # Create a list to store the table's data
    table_data2 = []
    # Iterate through the table rows and columns
    for row in range(min_row, max_row + 1):
        row_data = []
        for col in range(min_col, max_col + 1):
            cell_value = awread.cell(row=row, column=col).value
            row_data.append(cell_value)
        table_data2.append(row_data)
    studentmarks = pd.DataFrame(table_data2[1:], columns=table_data2[0])

    no_questions=qno_mm_co_btl.shape[1]-1


    awwrite.merge_cells(f'B1:{get_column_letter(qno_mm_co_btl.shape[1]+2)}1')
    awwrite['B1'] = f'{awread.title}'
    awwrite['B1'].font = Font(bold=True)
    for c, header in enumerate(qno_mm_co_btl.columns, start=2):
        awwrite.cell(row=2, column=c, value=header)
        awwrite.cell(row=2, column=c).font = Font(bold=True)
    # Write the DataFrame to the worksheet
    for r, row in enumerate(qno_mm_co_btl.values, start=3):
        for c, value in enumerate(row, start=2):
            awwrite.cell(row=r, column=c, value=value)
            if c==2:
                awwrite.cell(row=r, column=c).font = Font(bold=True)
                

    table_range= f'B2:{get_column_letter(qno_mm_co_btl.shape[1]+1)}{qno_mm_co_btl.shape[0]+2}'
    tab = Table(displayName=f"Input_CODetails_Table_{awread.title}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{table_colour}", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    awwrite.add_table(tab)


    awwrite.merge_cells(f'C9:{get_column_letter(qno_mm_co_btl.shape[1]+2)}9')
    awwrite['C9'] = f'Marks Obtained'
    awwrite['C9'].font = Font(bold=True)
    for c, header in enumerate(studentmarks.columns, start=1):
        awwrite.cell(row=10, column=c, value=header)
        awwrite.cell(row=10, column=c).font = Font(bold=True)
    # Write the DataFrame to the worksheet
    for r, row in enumerate(studentmarks.values, start=11):
        for c, value in enumerate(row, start=1):
            awwrite.cell(row=r, column=c, value=value)

    table_range= f'A10:{get_column_letter(studentmarks.shape[1])}{studentmarks.shape[0]+10}'
    tab = Table(displayName=f"Input_StudentMarks_Table_{awread.title}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{table_colour}", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    awwrite.add_table(tab)


    #print(qno_mm_co_btl.iloc[3,1:])
    c=list(qno_mm_co_btl.iloc[3,1:])
    c.insert(0,'Name')
    c.insert(0,'Roll No')
    studentmarks.columns = c

    qno_mm_co_btl = qno_mm_co_btl.transpose()
    qno_mm_co_btl.columns = ["Max Marks","Threshold","CO","Final CO","BTL"]
    qno_mm_co_btl = qno_mm_co_btl.drop(qno_mm_co_btl.index[0])
    qno_mm_co_btl = qno_mm_co_btl.groupby(['Final CO'])[["Max Marks","Threshold"]].sum().transpose()

    awwrite.merge_cells(f'{get_column_letter(no_questions+2+1+1)}4:{get_column_letter(no_questions+2+1+1+qno_mm_co_btl.shape[1])}4')
    awwrite[f'{get_column_letter(no_questions+2+1+1)}4'] = f'CO Details'
    awwrite[f'{get_column_letter(no_questions+2+1+1)}4'].font = Font(bold=True)
    
    awwrite[f"{get_column_letter(no_questions+1+2+1)}5"] = "CO"
    awwrite[f"{get_column_letter(no_questions+1+2+1)}5"].font = Font(bold=True)

    awwrite[f"{get_column_letter(no_questions+1+2+1)}6"] = "Max Marks"
    awwrite[f"{get_column_letter(no_questions+1+2+1)}6"].font = Font(bold=True)

    awwrite[f"{get_column_letter(no_questions+1+2+1)}7"] = "Threshold"
    awwrite[f"{get_column_letter(no_questions+1+2+1)}7"].font = Font(bold=True)


    # Write the header
    for c, header in enumerate(qno_mm_co_btl.columns, start=2+no_questions+3):
        awwrite.cell(row=5, column=c, value=header)

    # Write the DataFrame to the worksheet
    for r, row in enumerate(qno_mm_co_btl.values, start=6):
        for c, value in enumerate(row, start=2+no_questions+3):
            awwrite.cell(row=r, column=c, value=value)

    #Make table
    table_range= f'{get_column_letter(3+no_questions+2)}5:{get_column_letter(2+no_questions+2+qno_mm_co_btl.shape[1])}{7}'
    tab = Table(displayName=f"CummulativeCO_{str(awread.title)}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{table_colour+1}", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    awwrite.add_table(tab)



    name = studentmarks['Name']
    rollno = studentmarks['Roll No']
    studentmarks = studentmarks.drop(['Name','Roll No'], axis=1)
    #groupby index
    studentmarks =studentmarks.transpose()
    studentmarks = studentmarks.groupby(studentmarks.index).sum()
    #add name and rollno as the first two columns
    studentmarks = studentmarks.transpose()
    # studentmarks.insert(0, "Name", name)
    # studentmarks.insert(1, "Roll No", rollno)
    
    # Write the header
    for c, header in enumerate(studentmarks.columns, start=2+no_questions+3):
        awwrite.cell(row=10, column=c, value=header)

    # Write the DataFrame to the worksheet
    for r, row in enumerate(studentmarks.values, start=11):
        for c, value in enumerate(row, start=2+no_questions+3):
            awwrite.cell(row=r, column=c, value=value)

    # #make a table
    table_range= f'{get_column_letter(2+no_questions+3)}10:{get_column_letter(2+no_questions+2+studentmarks.shape[1])}{studentmarks.shape[0]+10}'
    tab = Table(displayName=f"CummulativeStudent_"+str(awread.title), ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{table_colour+1}", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    awwrite.add_table(tab)



    return awwrite

#Read all CO tables and calculate component wise CO marks
import openpyxl
import pandas as pd

def read_tables_in_excel_file(wb, component):
    
    # Load the workbook
    #wb = openpyxl.load_workbook(file_path, data_only=True)

    # Create an empty dictionary to store the DataFrames
    dfsCO = {}
    df_finalCO = pd.DataFrame()

    dfsStudent = {}
    df_finalStudent = pd.DataFrame()

    # Iterate through the sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Find all the tables in the sheet
        for table in sheet.tables.values():
            table_name = table.displayName

            # Check if the table name contains "Cummulative CO"
            if "CummulativeCO" in table_name and table_name.split("_")[-1] == component:
                print(f"Reading Table: {table_name}")

                # Get table dimensions
                table_range = table.ref
                min_col, min_row, max_col, max_row = range_boundaries(table_range)

                # Create a list to store the table's data
                table_data = []

                # Iterate through the table rows and columns
                for row in range(min_row, max_row + 1):
                    row_data = []
                    for col in range(min_col, max_col + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        row_data.append(cell_value)
                    table_data.append(row_data)

                # Convert the table's data to a DataFrame and store it in the dictionary with the table name as the key
                table_df = pd.DataFrame(table_data[1:], columns=table_data[0])
                
                dfsCO[table_name] = table_df
                df_finalCO = pd.concat([df_finalCO, table_df], axis=1)
        

            
        #Find all the tables in the sheet
        for table in sheet.tables.values():
            table_name = table.displayName

            # Check if the table name contains "Cummulative CO"
            if "CummulativeStudent" in table_name and table_name.split("_")[-1] == component:
                print(f"Reading Table: {table_name}")

                # Get table dimensions
                table_range = table.ref
                min_col, min_row, max_col, max_row = range_boundaries(table_range)

                # Create a list to store the table's data
                table_data = []

                # Iterate through the table rows and columns
                for row in range(min_row, max_row + 1):
                    row_data = []
                    for col in range(min_col, max_col + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        row_data.append(cell_value)
                    table_data.append(row_data)

                # Convert the table's data to a DataFrame and store it in the dictionary with the table name as the key
                table_df = pd.DataFrame(table_data[1:], columns=table_data[0])
                #table_df = table_df.drop(['Name','Roll No'], axis=1)
                dfsStudent[table_name] = table_df
                df_finalStudent = pd.concat([df_finalStudent, table_df], axis=1)

    dfs_finalCO_combined = df_finalCO.groupby(df_finalCO.columns, axis=1).sum()
    dfsCO["Final_Combined_CO"] = dfs_finalCO_combined

    dfs_finalStudent_combined = df_finalStudent.groupby(df_finalStudent.columns, axis=1).sum()
    dfsStudent["Final_Combined_Student"] = dfs_finalStudent_combined

    # Calculate the 50% threshold for each CO
   
    thresholds = dfs_finalCO_combined.loc[1]
    
    # Create a boolean DataFrame with True if the student's mark is greater than 50% of the maximum value in the corresponding CO, False otherwise
    above_50_percent = dfs_finalStudent_combined.gt(thresholds)

    # Calculate the number of students with more than 50% of marks for a given CO
    num_students_above_50_percent = above_50_percent.sum()

    # Calculate the CO %
    co_percentage = num_students_above_50_percent

    # Calculate the Attainment %
    attainment_percentage = co_percentage / len(dfs_finalStudent_combined) * 100

    # Create a new DataFrame with CO % and Attainment %
    result_df = pd.DataFrame({'CO':dfs_finalStudent_combined.columns,'CO %': co_percentage, 'Total Students':len(dfs_finalStudent_combined),f'{component}_Attainment %': attainment_percentage})

    return dfsCO, dfsStudent, result_df

def create_template(aw,data,COPOTable_df,final_table, result):
    
    #fill COPOTable_df with 0 for missing values
    COPOTable_df = COPOTable_df.fillna(0)


    aw.merge_cells('A1:A4')
    aw['A1'] = 'Course Outcome'
    aw['A1'].font = Font(bold=True)

    aw.merge_cells('B1:C1')
    aw['B1'] = 'Mapping with Program'
    aw['B1'].font = Font(bold=True)

    aw.merge_cells('D1:K1')
    aw['D1'] = 'Attainment % in'
    aw['D1'].font = Font(bold=True)

    aw.merge_cells('B2:B4')
    aw['B2'] = 'POs & PSOs'
    aw['B2'].font = Font(bold=True)

    aw["C2"]="Level of Mapping"
    aw["C2"].font = Font(bold=True)

    aw.merge_cells('C3:C4')
    aw['C3'] = 'Affinity'
    aw['C3'].font = Font(bold=True)

    aw.merge_cells('D2:H2')
    aw['D2'] = 'Direct'
    aw['D2'].font = Font(bold=True)

    aw.merge_cells('I2:J2')
    aw['I2'] = 'Indirect'
    aw['I2'].font = Font(bold=True)

    aw.merge_cells('K2:K3')
    aw['K2'] = 'Final Weighted CO Attainment (80% Direct + 20% Indirect)'
    aw['K2'].font = Font(bold=True)

    aw.merge_cells('D3:E3')
    aw['D3'] = 'University(SEE)'
    aw['D3'].font = Font(bold=True)

    aw.merge_cells('F3:G3')
    aw['F3'] = 'Internal(CIE)'
    aw['F3'].font = Font(bold=True)

    aw.merge_cells('H3:H4')
    aw['H3'] = 'Weighted Level of Attainment (University + IA)'
    aw['H3'].font = Font(bold=True)

    aw["D4"]="Attainment"
    aw["D4"].font = Font(bold=True)

    aw["E4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["E4"].font = Font(bold=True)

    aw["F4"]="Attainment"
    aw["F4"].font = Font(bold=True)

    aw["G4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["G4"].font = Font(bold=True)

    aw.merge_cells('I3:I4')
    aw["I3"]="Attainment"
    aw["I3"].font = Font(bold=True)

    aw.merge_cells('J3:J4')
    aw['J3']="Level Of Attainment"
    aw["J3"].font = Font(bold=True)

    aw["K4"]="Level of Attainment"
    aw["K4"].font = Font(bold=True)

    #Set column width for A to 17.22
    aw.column_dimensions['A'].width = 17.22
    aw.column_dimensions['B'].width = 9.33
    aw.column_dimensions['C'].width = 15.56
    aw.column_dimensions['D'].width = 10.33
    aw.column_dimensions['E'].width = 14.11
    aw.column_dimensions['F'].width = 10.33
    aw.column_dimensions['G'].width = 14.11
    aw.column_dimensions['H'].width = 20.67
    aw.column_dimensions['I'].width = 10.33
    aw.column_dimensions['J'].width = 18.11
    aw.column_dimensions['K'].width = 22.78
    #center align the text in the cells
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            #set color of the cells to blue
            cell.fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    #set color of the cells to green
    aw["C3"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw["K4"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')

    start=4
    interval=16

    for i in range(1, (data["Number_of_COs"]+1)):
        
        start+=1
        aw.merge_cells(start_row=start, start_column=1, end_row=start+interval, end_column=1)
        aw.cell(row=start, column=1).value = "CO"+str(i)
        aw.cell(row=start, column=1).font = Font(bold=True)
        aw.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i%2==0:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
        else:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')

        index=1
        for j in range(start, start+interval+1):
            #print out COPO mapping
            aw.cell(row=j, column=2).value = COPOTable_df.columns[index]
            aw.cell(row=j, column=3).value = COPOTable_df.iloc[i-1, index]
            if index%2==0:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
            else:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            index+=1
        
        

        for k in range(4,12):
            aw.merge_cells(start_row=start, start_column=k, end_row=start+interval, end_column=k)
            aw.cell(row=start, column=k).value = final_table.iloc[i-1, k-4]
            aw.cell(row=start, column=k).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if k%2==0:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            else:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='dce6f1', end_color='dce6f1', fill_type='solid')


            aw.cell(row=start, column=k).border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
            
    



        start=start+interval
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
    
    #find out the last row in the excel sheet
    current_row = aw.max_row+4
    aw.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=17+2)
    aw.cell(row=current_row, column=2).value = "Weighted PO/PSO Attainment Contribution"
    aw.cell(row=current_row, column=2).font = Font(bold=True)
    aw.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    

    # Print the resulting DataFrame
    # Write the header
    for c, header in enumerate(result.columns, start=2):
        aw.cell(row=current_row+1, column=c, value=header)

    # Write the DataFrame to the worksheet
    for r, row in enumerate(result.values, start=current_row+2):
        for c, value in enumerate(row, start=2):
            aw.cell(row=r, column=c, value=value)

    #create a table
    tab = Table(displayName="WeightedPO", ref="B"+str(current_row+1)+":S"+str(current_row+1+len(result)))
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    aw.add_table(tab)

    for row in aw.iter_rows(min_row=current_row, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row: 
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    return aw

def write_course_level_attainment(aw, data, COPOTable_df, result_df_I, result_df_E, Indirect_Attainment_df):
    def con(x):
        x = float(x)
        if x > 60 and x <= 100:
            return 3
        elif x > 40 and x <= 60:
            return 2
        elif x > 20 and x <= 40:
            return 1
        else:
            return 0

    final_table = result_df_E.transpose()
    final_table.drop(final_table.index[0:3], inplace=True)
    final_table.loc['E_LOA'] = final_table.iloc[0, :].apply(con)

    final_table = pd.concat([final_table, result_df_I.transpose().drop(result_df_I.transpose().index[0:3])]).fillna(0)
    final_table.loc['I_LOA'] = final_table.iloc[2, :].apply(con)
    final_table.loc['Weighted LOA'] = (data["Internal"] / 100) * final_table.loc['I_LOA'] + (data["External"] / 100) * final_table.loc['E_LOA']

    indirect_attainment = pd.DataFrame([Indirect_Attainment_df.transpose().iloc[1].values], columns=final_table.columns, index=["Indirect Attainment"])
    # Concatenate final_table and indirect_attainment
    final_table = pd.concat([final_table, indirect_attainment])
    final_table.loc["Indirect LOA"] = final_table.iloc[5, :].apply(con)
    final_table.loc["Final LOA"] = (final_table.loc["Weighted LOA"] * (data["Direct"])/100) + (final_table.loc["Indirect LOA"] * (data["Indirect"])/100)
    final_table=final_table.transpose()

    weightedPO = COPOTable_df.copy()
    weightedPO.iloc[:, 1:] = 0

    # Get Final LOA values from the second table
    final_loa = final_table['Final LOA']

    # Initialize an empty DataFrame with the same shape as weightedPO to store the result
    result = pd.DataFrame(index=weightedPO.index, columns=weightedPO.columns)

    # Iterate through each cell in the weightedPO table and multiply with the corresponding Final LOA
    for i, row in COPOTable_df.iterrows():
        for j, value in enumerate(row):
            if j > 0 and value is not None:  # Start from the second column (index 1) and skip None values
                result.iloc[i, j] = value * final_loa[i]

    # Set the first column ('COs\POs') in the result DataFrame to match the original weightedPO table
    result['COs\\POs'] = COPOTable_df['COs\\POs']
    

    # Calculate the sum of the values in each column of the result DataFrame, ignoring the first row ('COs\POs')
    sums = result.iloc[:, 1:].sum(axis=0)

    # Calculate the sum of the values in each row of the COPOTable_df DataFrame, ignoring the first column ('COs\POs')
    total_sums = COPOTable_df.iloc[:, 1:].sum(axis=0)

    # Check if the total sum is greater than 0, and divide the sums by the total_sums element-wise, handling division by zero
    ratio = []
    for i, total_sum in enumerate(total_sums):
        if total_sum > 0:
            ratio.append(sums[i] / total_sum)
        else:
            ratio.append(0)
    # Create a new DataFrame with the same columns as result and a single row called 'Final ratio'
    final_ratio_df = pd.DataFrame([ratio], columns=result.columns[1:])

    # Concatenate the result DataFrame and the final_ratio_df
    result= pd.concat([result, final_ratio_df])
    result.iloc[-1, 0] = data["Subject_Code"]

    aw = create_template(aw, data, COPOTable_df, final_table, result)
    
    return aw, final_table

def printout_template(aw, data):
    aw.merge_cells("D1:D3")
    aw["D1"]="Course"
    aw["D1"].font = Font(bold=True)
    aw["D1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=3, min_col=4, max_col=4):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            

    aw.merge_cells("E1:E3")
    aw["E1"]="COs"
    aw["E1"].font = Font(bold=True)
    aw["E1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=3, min_col=5, max_col=5):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells("F1:G1")
    aw["F1"]="End Semester Examination"
    aw["F1"].font = Font(bold=True)
    aw["F1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=1, min_col=6, max_col=7):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw.merge_cells("F2:G2")
    aw["F2"]="(SEE)*"
    aw["F2"].font = Font(bold=True)
    aw["F2"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=2, max_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw["F3"]="Attainment"
    aw["F3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["F3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["G3"]="Level"
    aw["G3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["G3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells("H1:I1")
    aw["H1"]="Internal Examination"
    aw["H1"].font = Font(bold=True)
    aw["H1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=1, min_col=8, max_col=9):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw.merge_cells("H2:I2")
    aw["H2"]="(CIE)*"
    aw["H2"].font = Font(bold=True)
    aw["H2"].alignment = Alignment(horizontal='center', vertical='center')

    aw["H3"]="Attainment"
    aw["H3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["H3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["I3"]="Level"
    aw["I3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["I3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    

    aw.merge_cells("J1:K1")
    aw["J1"]="Direct"
    aw["J1"].font = Font(bold=True)
    aw["J1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=1, min_col=10, max_col=11):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells("J2:K2")
    aw["J2"]=f"{data['Internal']}% of CIE + {data['External']}% of SEE"
    aw["J2"].font = Font(bold=True)
    aw["J2"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=2, max_row=2, min_col=10, max_col=11):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw["J3"]="Attainment"
    aw["J3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["J3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["K3"]="Level"
    aw["K3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["K3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells("L1:M2")
    aw["L1"]="Indirect"
    aw["L1"].font = Font(bold=True)
    aw["L1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=2, min_col=12, max_col=13):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw["L3"]="Attainment"
    aw["L3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["L3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["M3"]="Level"
    aw["M3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["M3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells("N1:O1")
    aw["N1"]="Total Course Attainment"
    aw["N1"].font = Font(bold=True)
    aw["N1"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=1, max_row=1, min_col=14, max_col=15):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells("N2:O2")
    aw["N2"]=f"{data['Direct']}% of Direct + {data['Indirect']}% of Indirect"
    aw["N2"].font = Font(bold=True)
    aw["N2"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in aw.iter_rows(min_row=2, max_row=2, min_col=14, max_col=15):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw["N3"]="Attainment"
    aw["N3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["N3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["O3"]="Level"
    aw["O3"].alignment = Alignment(horizontal='center', vertical='center')
    aw["O3"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw["P1"]="Target"
    aw["P1"].font = Font(bold=True)
    aw["P1"].alignment = Alignment(horizontal='center', vertical='center')
    aw["P1"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            
    aw["P2"]="(%)"
    aw["P2"].font = Font(bold=True)
    aw["P2"].alignment = Alignment(horizontal='center', vertical='center')
    aw["P2"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    aw["Q1"]="Attainment"
    aw["Q1"].font = Font(bold=True)
    aw["Q1"].alignment = Alignment(horizontal='center', vertical='center')
    aw["Q1"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    aw["Q2"]="Yes/No"
    aw["Q2"].font = Font(bold=True)
    aw["Q2"].alignment = Alignment(horizontal='center', vertical='center')
    aw["Q2"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    
    aw.column_dimensions['D'].width = 8.43
    aw.column_dimensions['E'].width = 8.43
    aw.column_dimensions['F'].width = 12
    aw.column_dimensions['G'].width = 12
    aw.column_dimensions['H'].width = 12
    aw.column_dimensions['I'].width = 12
    aw.column_dimensions['J'].width = 12
    aw.column_dimensions['K'].width = 12
    aw.column_dimensions['L'].width = 12
    aw.column_dimensions['M'].width = 8.43
    aw.column_dimensions['N'].width = 20
    aw.column_dimensions['O'].width = 8.43
    aw.column_dimensions['P'].width = 8.43
    aw.column_dimensions['Q'].width = 10

    
    for row in aw.iter_cols(min_row=3, max_row=3, min_col=5, max_col=aw.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='8db4e2', end_color='8db4e2', fill_type='solid')

def printout(aw, data, COPOTable_df, final_table,result_df_I, result_df_E, Indirect_Attainment_df):

    aw.cell(row=1, column=1).value = "Headings"
    aw.cell(row=1, column=1).font = Font(bold=True)
    aw.cell(row=1, column=2).value = "Values"
    aw.cell(row=1, column=2).font = Font(bold=True)

    current_row=2
    head=["Batch","Subject_Code","Subject_Name","Academic_year","Semester"]
    for key in data.keys():
        aw.cell(row=current_row, column=1).value = key+" :"
        aw.cell(row=current_row, column=2).value = data[key]
        aw.cell(row=current_row, column=1).font = Font(bold=True)
        aw.cell(row=current_row, column=2).font = Font(bold=True)
        aw.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
        current_row+=1

    #add a table
    tab = Table(displayName="Inputinfo", ref="A1:B"+str(current_row-1))
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    aw.add_table(tab)


    adjust_width(aw)
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row: 
            cell.alignment = Alignment(horizontal='center', vertical='center')

    printout_template(aw,data)

    #merge D4 to number of COs
    aw.merge_cells(f"D4:D{3+data['Number_of_COs']}")
    #write the course name horizontally
    aw["D4"]=data["Subject_Name"]
    aw["D4"].font = Font(bold=True)
    aw["D4"].alignment = Alignment(horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    aw["D4"].fill = PatternFill(start_color='1ed760', end_color='1ed760', fill_type='solid')

    for i in range(data["Number_of_COs"]):
        aw[f"E{4+i}"]=f"CO{i+1}"
        aw[f"E{4+i}"].font = Font(bold=True)
        aw[f"E{4+i}"].alignment = Alignment(horizontal='center', vertical='center')
        if i%2==0:
            aw[f"E{4+i}"].fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        aw[f"E{4+i}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        
    

    Prinout_COPO_df=final_table.copy()
    Prinout_COPO_df.insert(4,"Direct Attainment", np.zeros(len(Prinout_COPO_df)))
    Prinout_COPO_df.insert(5,"Direct_LOA", np.zeros(len(Prinout_COPO_df)))
    Prinout_COPO_df.drop(columns=["Weighted LOA","Final LOA"], inplace=True)
    #add column to end of table
    Prinout_COPO_df.insert(len(Prinout_COPO_df.columns),"Total Course Attainment", np.zeros(len(Prinout_COPO_df)))
    Prinout_COPO_df.insert(len(Prinout_COPO_df.columns),"Total Course LOA", np.zeros(len(Prinout_COPO_df)))
    Prinout_COPO_df.insert(len(Prinout_COPO_df.columns),"Target", data["Default threshold %"])
    Prinout_COPO_df.insert(len(Prinout_COPO_df.columns),"Attainment", np.zeros(len(Prinout_COPO_df)))

    #Prinout_COPO_df=Prinout_COPO_df.transpose()
    for rowno in range(len(Prinout_COPO_df)):
        Prinout_COPO_df.iloc[rowno,4]=(Prinout_COPO_df.iloc[rowno,0]*(data["External"]/100))+((data["Internal"]/100)*Prinout_COPO_df.iloc[rowno,2])
        Prinout_COPO_df.iloc[rowno,5]=(Prinout_COPO_df.iloc[rowno,1]*(data["External"]/100))+((data["Internal"]/100)*Prinout_COPO_df.iloc[rowno,3])
        Prinout_COPO_df.iloc[rowno,8]=(Prinout_COPO_df.iloc[rowno,4]*(data["Direct"]/100))+((data["Indirect"]/100)*Prinout_COPO_df.iloc[rowno,6])
        Prinout_COPO_df.iloc[rowno,9]=(Prinout_COPO_df.iloc[rowno,5]*(data["Direct"]/100))+((data["Indirect"]/100)*Prinout_COPO_df.iloc[rowno,7])
        Prinout_COPO_df.iloc[rowno, 11] = '=IF($N$' + str(4 + rowno) + '>= $P$' + str(4 + rowno) + ', "Yes", "No")'


    #print it to excel
    startrow=4
    startcol=6
    for rowno in range(len(Prinout_COPO_df)):
        for colno in range(len(Prinout_COPO_df.columns)):
            aw.cell(row=startrow+rowno, column=startcol+colno).value = Prinout_COPO_df.iloc[rowno,colno]
            aw.cell(row=startrow+rowno, column=startcol+colno).alignment = Alignment(horizontal='center', vertical='center')
            aw.cell(row=startrow+rowno, column=startcol+colno).border = Border(left=Side(border_style='thin', color='000000'),
                                                                                right=Side(border_style='thin', color='000000'),
                                                                                top=Side(border_style='thin', color='000000'),
                                                                                bottom=Side(border_style='thin', color='000000'))
            if colno==len(Prinout_COPO_df.columns)-1:
                aw.cell(row=startrow+rowno, column=startcol+colno).font = Font(bold=True)

            if colno==1 or colno==3:
                aw.cell(row=startrow+rowno, column=startcol+colno).fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
                #set the font to red
                aw.cell(row=startrow+rowno, column=startcol+colno).font = Font(color='ff0000')

    

            
    
    return aw

#write the DataFrames from dfsCO, dfsStudent and co_attainment_df to the Excel file
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

def write_tables_to_excel(aw, dfsCO, dfsStudent, co_attainment_df, component, tc):
    # Create a new workbook

    # Set the initial column position for the first table
    current_col = 1
    table_colour = tc

    # Iterate through the DataFrames from dfsCO
    for table_name, df in dfsCO.items():
        table_colour += 2
        # Write the table name and merge the cells
        aw.cell(row=1, column=current_col, value=table_name)
        aw.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(df.columns) - 1)

        # Write the DataFrame to the worksheet
        for r, row in enumerate(df.values, start=3):
            for c, value in enumerate(row, start=current_col):
                aw.cell(row=r, column=c, value=value)

        # Write the header
        for c, header in enumerate(df.columns, start=current_col):
            aw.cell(row=2, column=c, value=header)

        # Create a table for the DataFrame
        table_range = get_column_letter(current_col) + "2:" + get_column_letter(current_col + len(df.columns) - 1) + str(len(df) + 2)
        table = Table(displayName=table_name+str(component), ref=table_range)
        style = TableStyleInfo(name=f"TableStyleMedium{table_colour}", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        aw.add_table(table)

        # Move the current_col to the next position, leaving a two-column gap
        current_col += len(df.columns) + 2

    # Reset the initial column position for the dfsStudent tables
    current_col = 1
    # Calculate the row number to start writing the dfsStudent DataFrames
    start_row = 5
    table_colour=tc
    # Iterate through the DataFrames from dfsStudent
    for table_name, df in dfsStudent.items():
        table_colour += 2
        
        # Write the DataFrame to the worksheet
        for r, row in enumerate(df.values, start=start_row + 2):
            for c, value in enumerate(row, start=current_col):
                aw.cell(row=r, column=c, value=value)

        # Write the header
        for c, header in enumerate(df.columns, start=current_col):
            aw.cell(row=start_row + 1, column=c, value=header)

        # Create a table for the DataFrame
        table_range = get_column_letter(current_col) + str(start_row + 1) + ":" + get_column_letter(current_col + len(df.columns) - 1) + str(len(df)+start_row+1)
        table = Table(displayName=table_name+str(component), ref=table_range)
        style = TableStyleInfo(name=f"TableStyleMedium{table_colour}", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        aw.add_table(table)

        current_col += len(df.columns) + 2

    start_row = aw.max_row + 3
    start_col = aw.max_column - len(dfsStudent["Final_Combined_Student"].columns) + 1

    #transposing the dataframe with the row index as the column headers
    co_attainment_df = co_attainment_df.T


    # Write the DataFrame to the worksheet
    for r, (_, row) in enumerate(co_attainment_df.iterrows(), start=start_row):
        for c, value in enumerate(row, start=start_col):
            aw.cell(row=r, column=c, value=value)

    # Write the index (CO and Attainment %) as row headers
    for r, index_val in enumerate(co_attainment_df.index, start=start_row):
        aw.cell(row=r, column=start_col - 1, value=index_val)

   

    # Create a table for the DataFrame
    table_range = f"{get_column_letter(start_col-1)}{start_row}:{get_column_letter(start_col + len(co_attainment_df.columns) - 1)}{start_row + len(co_attainment_df) - 1}"
    table = Table(displayName=f"CO_Attainment_{component}", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    aw.add_table(table)

                
    return aw

def driver_part2(path):

    # path = 'C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\Part 1\\v16.xlsx'


    wbread = load_workbook(path, data_only=True)

    
    wsread=wbread["Input Details"]
    data={}
    inputTable=wsread.tables['Input_Details']
    # Iterate through the rows in the input table
    table_range=inputTable.ref
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    for row in range(min_row, max_row+1):
        data[wsread.cell(row=row, column=1).value]=wsread.cell(row=row, column=2).value


    COPOTable = wsread.tables['CO_PO']
    #iterate and save it in dataframes
    table_range=COPOTable.ref
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    rowdata=[]
    for row in range(min_row, max_row+1):
        rowdata.append([])
        for col in range(min_col, max_col+1):
            rowdata[-1].append(wsread.cell(row=row, column=col).value)
    COPOTable_df = pd.DataFrame(rowdata[1:], columns=rowdata[0])

    Indirect_Attainment = wsread.tables["Indirect_CO_Assessment"]
    table_range=Indirect_Attainment.ref
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    rowdata=[]
    for row in range(min_row, max_row+1):
        rowdata.append([])
        for col in range(min_col, max_col+1):
            rowdata[-1].append(wsread.cell(row=row, column=col).value)
    Indirect_Attainment_df = pd.DataFrame(rowdata[1:], columns=rowdata[0])



    time.sleep(2)                                                                            #sleep for 2 seconds
    wbwrite = Workbook()                                                                 #create workbook
    wbwrite.remove(wbwrite.active)
    wbwrite = create_sheets(path, wbwrite)
    #get number of sheets
    num_components = len(wbwrite.sheetnames)

    i=1
    for sheet in wbwrite.sheetnames:
        i+=2
        wswrite=wbwrite[sheet]
        wsread=wbread[sheet[3:]]
        wswrite = CummulativeCO(wswrite, wsread, i)
        

    wbwrite.create_sheet("Internal Component")
    wswrite=wbwrite["Internal Component"]
    dfsCO, dfsStudent, result_df_I = read_tables_in_excel_file(wbwrite, "I")
    wswrite = write_tables_to_excel(wswrite, dfsCO, dfsStudent, result_df_I, "I",2)

    wbwrite.create_sheet("External Component")
    wswrite=wbwrite["External Component"]
    dfsCO, dfsStudent, result_df_E = read_tables_in_excel_file(wbwrite, "E")
    wswrite = write_tables_to_excel(wswrite, dfsCO, dfsStudent, result_df_E, "E",num_components*2)

    #iterate through all the sheets and set the width of the columns
    for s in wbwrite.worksheets:
        adjust_width(s)
        #center alignment for all the cells
        for row in s.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')  

    wbwrite.create_sheet("Course Level Attainment")
    wswrite=wbwrite["Course Level Attainment"]
    wswrite, final_table=write_course_level_attainment(wswrite, data, COPOTable_df, result_df_I, result_df_E, Indirect_Attainment_df)

    wbwrite.create_sheet("Printout")
    wswrite=wbwrite["Printout"]
    wswrite=printout(wswrite, data, COPOTable_df,final_table, result_df_I, result_df_E, Indirect_Attainment_df)
    
    
    #name=path.split()
    print(path)
    wbwrite.save(f"Calculated_{data['Batch']}_{data['Subject_Code']}_{data['Subject_Name']}.xlsx")